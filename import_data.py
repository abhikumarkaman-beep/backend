# KrishiConnect AI - Excel Data Import Script
# Imports Book1_final.xlsx, disease_product_mapping.xlsx into SQLite

import sqlite3
import openpyxl
import os
import sys
from config import Config

sys.stdout.reconfigure(encoding='utf-8')

# Paths
BASE_DIR = os.path.dirname(__file__)
DATA_DIR = os.path.join(BASE_DIR, 'data')
DB_PATH = Config.DATABASE_PATH

BOOK1_PATH = os.path.join(DATA_DIR, 'Book1_final.xlsx')
DISEASE_MAP_PATH = os.path.join(DATA_DIR, 'disease_product_mapping.xlsx')

def import_districts_and_crops():
    """Import Book1_final.xlsx → districts + district_crops tables"""
    
    print("=" * 50)
    print("Importing Book1_final.xlsx...")
    
    wb = openpyxl.load_workbook(BOOK1_PATH)
    ws = wb.active
    
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    
    # Track unique districts (to avoid duplicates)
    district_ids = {}  # "State|District" → id
    
    districts_added = 0
    crops_added = 0
    skipped = 0
    
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        state = row[0].value
        district = row[1].value
        crop = row[2].value
        season = row[3].value
        language = row[4].value
        lat = row[5].value
        lon = row[6].value
        
        if not all([state, district, crop, season, language, lat, lon]):
            skipped += 1
            continue
        
        # Insert district (if not already inserted)
        key = f"{state}|{district}"
        if key not in district_ids:
            try:
                cursor.execute("""
                    INSERT OR IGNORE INTO districts (state, district, language, latitude, longitude)
                    VALUES (?, ?, ?, ?, ?)
                """, (state, district, language, lat, lon))
                
                if cursor.lastrowid:
                    district_ids[key] = cursor.lastrowid
                    districts_added += 1
                else:
                    # Already existed, get its ID
                    cursor.execute(
                        "SELECT id FROM districts WHERE state=? AND district=?",
                        (state, district)
                    )
                    result = cursor.fetchone()
                    if result:
                        district_ids[key] = result[0]
            except Exception as e:
                # Get existing ID
                cursor.execute(
                    "SELECT id FROM districts WHERE state=? AND district=?",
                    (state, district)
                )
                result = cursor.fetchone()
                if result:
                    district_ids[key] = result[0]
        
        # Insert district-crop mapping
        district_id = district_ids.get(key)
        if district_id:
            try:
                cursor.execute("""
                    INSERT OR IGNORE INTO district_crops (district_id, crop, season)
                    VALUES (?, ?, ?)
                """, (district_id, crop, season))
                if cursor.rowcount > 0:
                    crops_added += 1
            except:
                pass
    
    conn.commit()
    print(f"  Districts added: {districts_added}")
    print(f"  Crop mappings added: {crops_added}")
    print(f"  Skipped rows: {skipped}")
    
    conn.close()
    wb.close()


def import_disease_product_mapping():
    """Import disease_product_mapping.xlsx → disease_product_map table"""
    
    print("\n" + "=" * 50)
    print("Importing disease_product_mapping.xlsx...")
    
    wb = openpyxl.load_workbook(DISEASE_MAP_PATH)
    ws = wb.active
    
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    
    added = 0
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        disease_name = row[0].value
        if not disease_name:
            continue
        
        is_viral = 1 if str(row[11].value).upper() == 'TRUE' else 0
        
        try:
            cursor.execute("""
                INSERT OR REPLACE INTO disease_product_map 
                (disease_name, affected_crops, disease_type, product_category,
                 primary_product, primary_chemical, primary_dosage, primary_application,
                 secondary_product, secondary_chemical, secondary_dosage,
                 is_viral, vector_insect, notes)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                disease_name, row[1].value, row[2].value, row[3].value,
                row[4].value, row[5].value, row[6].value, row[7].value,
                row[8].value, row[9].value, row[10].value,
                is_viral, row[12].value, row[13].value
            ))
            added += 1
        except Exception as e:
            print(f"  Error with {disease_name}: {e}")
    
    conn.commit()
    print(f"  Disease-product mappings added: {added}")
    
    conn.close()
    wb.close()


def verify_import():
    """Verify all data was imported correctly"""
    
    print("\n" + "=" * 50)
    print("VERIFICATION:")
    
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    
    # Districts
    cursor.execute("SELECT COUNT(*) FROM districts")
    print(f"  Districts: {cursor.fetchone()[0]}")
    
    # States
    cursor.execute("SELECT COUNT(DISTINCT state) FROM districts")
    print(f"  States: {cursor.fetchone()[0]}")
    
    # Crops
    cursor.execute("SELECT COUNT(*) FROM district_crops")
    print(f"  District-Crop mappings: {cursor.fetchone()[0]}")
    
    cursor.execute("SELECT DISTINCT crop FROM district_crops ORDER BY crop")
    crops = [r[0] for r in cursor.fetchall()]
    print(f"  Unique crops: {crops}")
    
    # Seasons
    cursor.execute("SELECT season, COUNT(*) FROM district_crops GROUP BY season")
    for row in cursor.fetchall():
        print(f"    {row[0]}: {row[1]} entries")
    
    # Diseases
    cursor.execute("SELECT COUNT(*) FROM disease_product_map")
    print(f"  Disease-product mappings: {cursor.fetchone()[0]}")
    
    # Sample data
    cursor.execute("""
        SELECT d.state, d.district, dc.crop, dc.season, d.language, d.latitude, d.longitude
        FROM districts d JOIN district_crops dc ON d.id = dc.district_id
        LIMIT 5
    """)
    print("\n  Sample data:")
    for row in cursor.fetchall():
        print(f"    {row[0]} | {row[1]} | {row[2]} | {row[3]} | {row[4]} | ({row[5]}, {row[6]})")
    
    # Sample disease mapping
    cursor.execute("SELECT disease_name, primary_product, primary_dosage, is_viral FROM disease_product_map LIMIT 5")
    print("\n  Sample disease-product mapping:")
    for row in cursor.fetchall():
        viral = "🦠 VIRAL" if row[3] else ""
        print(f"    {row[0]} → {row[1]} ({row[2]}) {viral}")
    
    conn.close()


if __name__ == '__main__':
    print("KrishiConnect AI — Data Import")
    print("=" * 50)
    
    # First, initialize database
    from database import init_db
    init_db()
    
    # Import data
    import_districts_and_crops()
    import_disease_product_mapping()
    
    # Verify
    verify_import()
    
    print("\n" + "=" * 50)
    print("✅ All data imported successfully!")
    print("Database location:", DB_PATH)
